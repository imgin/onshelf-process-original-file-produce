# -*- coding: cp936 -*-
'''
python version:3.6
pip3 install xlrd
pip3 install xlwt
pip3 install xlutils
pip3 install openpyxl
pip3 install requests
pip3 install selenium
pip3 install pymongo
pip3 install python-dateutil
'''
import xlrd
import xlwt
from xlutils.copy import copy
from collections import OrderedDict
from urllib3 import encode_multipart_formdata
from uuid import uuid4
from pathlib import Path
import datetime
import os
import openpyxl
import time
import json
import requests
import re
import shutil
import zipfile
import selenium

def pimFileGenerate(environment,requiredNum,time_stamp,spuCodes,huohao):
    path = Path.cwd()
    excels = path.glob('*.xls*')
    modelCount=0
    colorNum = 3
    for excel in excels:
        filename = str(excel.name)
        if "pimԴ����.xls" in filename and "~$pimԴ����" not in filename:
            #print("hello world"+filename)
            print("�ҵ�pimԴ����"+"'"+filename+"'"+"���޸�������")
            modelCount = modelCount + 1
            data = openpyxl.load_workbook(filename)
            table = data['Դ����']
            #uu = datetime.datetime.now().strftime('%m%d')
            #huohao = random.randint(100, 999)
            #simplenum = random.randint(1, 99)
            a = 0
            #���ɻ���ROSSTEST305319,ROSSTEST305319-1
            GTIN_NUMBER = []
            while a < requiredNum*colorNum:
                GTIN_NUMBER.append(str(random.randint(100000000, 999999999)))
                a = a + 1
            print('��SheetԴ�������޸�ģ��')
            #nrows = table.nrows
            currColumn = 1
            currRow = 2
            maxColumnLen = table.max_column
            maxRowLen =requiredNum*colorNum+1
            #����colorNum�����ݽ����б�sample[][]
            num = 0
            sample = []
            while num < colorNum:
                sample.append([])
                currColumn = 1
                while currColumn <= maxColumnLen:
                    sample[num].append(str(table.cell(num+2,currColumn).value).strip())
                    currColumn = currColumn + 1
                num = num + 1
            #ճ����䵽colorNum*colorNum + 1��
            num = 1 #�ӵڶ�����ʼ
            currColumn = 1 
            while num < requiredNum:
                currColor = 0
                while currColor < colorNum:
                    currColumn = 1
                    while currColumn <= len(sample[currColor]):
                        table.cell(3*num + 2 + currColor,currColumn).value = sample[currColor][currColumn-1]
                        currColumn = currColumn + 1
                    currColor = currColor + 1
                num = num + 1
                
            currColumn = 1
            while currColumn <=  maxColumnLen:
                if table.cell(1,currColumn).value.strip() == "PDP_GROUPING":
                    while currRow <= maxRowLen:
                        table.cell(currRow,currColumn).value = spuCodes[int((currRow-2)/colorNum)]
                        currRow = currRow+1
                    currRow = 2
                    
                elif table.cell(1,currColumn).value.strip() == "GTIN_NUMBER":
                    while currRow <= maxRowLen:
                        table.cell(currRow,currColumn).value = "ROSSTS" + GTIN_NUMBER[currRow-2]
                        currRow = currRow+1
                    currRow = 2
                elif table.cell(1,currColumn).value.strip() == "COLLECTION_NUMBER":
                    while currRow <= maxRowLen:
                        table.cell(currRow,currColumn).value = huohao[int((currRow-2)/colorNum)]
                        currRow = currRow+1
                    currRow = 2
                elif table.cell(1,currColumn).value.strip() == "SHORT_SKU":
                    while currRow <= maxRowLen:
                        table.cell(currRow,currColumn).value = "ROSSTS" + GTIN_NUMBER[currRow-2]
                        currRow = currRow+1
                    currRow = 2
                currColumn = currColumn + 1


            resultPath = str(path) +'/' + time_stamp + '��' + str(requiredNum) + '��'
            if os.path.exists(resultPath):
                print('�Ѵ����ļ���'+resultPath +'����ɾ����һ���Ӻ�����')
            else:
                os.mkdir(resultPath)
            data.save(resultPath + '/pimԴ����'+ str(requiredNum) + '����'+time_stamp+'���.xlsx')
            return(createPlanExcel(spuCodes,environment,time_stamp,resultPath))

def createPlanExcel(spuCodes,environment,time_stamp,resultPath):
    path = Path.cwd()
    excels = path.glob('*.xls*')
    for excel in excels:
        filename = str(excel.name)
        print(filename)
        if 'WLQTESTHH001-1 - '+environment+'.xlsx' in filename and "~$WLQTESTHH001-1" not in filename:
            print("�ҵ��ƻ�ģ��"+"'"+filename+"'"+"���޸�������")
            # ��
            data = openpyxl.load_workbook(filename)
            table = data['Sheet1']
            # ������л���,ɫ�ŵļ�¼,
            num = 0
            while num < len(spuCodes):
                table.cell(num*3+2,1).value = spuCodes[num]
                table.cell(num*3+2,2).value = '2'
                table.cell(num*3+2,3).value = datetime.datetime.strptime('2020-2-25','%Y-%m-%d')
                table.cell(num*3+2,4).value = datetime.datetime.strptime('2020-3-28','%Y-%m-%d')
                table.cell(num*3+3,1).value = spuCodes[num]
                table.cell(num*3+3,2).value = '3'
                table.cell(num*3+3,3).value = datetime.datetime.strptime('2020-2-25','%Y-%m-%d')
                table.cell(num*3+3,4).value = datetime.datetime.strptime('2020-3-28','%Y-%m-%d')
                table.cell(num*3+4,1).value = spuCodes[num]
                table.cell(num*3+4,2).value = '4'
                table.cell(num*3+4,3).value = datetime.datetime.strptime('2020-2-25','%Y-%m-%d')
                table.cell(num*3+4,4).value = datetime.datetime.strptime('2020-3-28','%Y-%m-%d')
                num = num + 1
            data.save(resultPath+ '/WLQTESTHH001-1 - ' +environment + str(len(spuCodes)) + '����'+time_stamp+'���.xlsx')
    return('WLQTESTHH001-1 - ' +environment + str(len(spuCodes)) + '����'+time_stamp+'���.xlsx')
#�ϴ��ƻ�ģ��
def createSchedule(url,environment,token,shopcode,time_stamp,planFileName,requiredNum):
    url = url + '/schedule/createSchedule'
    '''with open("WLQTESTHH001-1.xls",'rb') as f:
        content = f.readlines()
       # print(content)
        content = ''.join(str(content)) #�˶δ�����content = open("WLQTESTHH001-1.xlsx", 'rb').read()������ͬ'''
    path=Path.cwd()
    resultPath = str(path) + '/' + time_stamp + '��' + str(requiredNum) + '��'
    files = OrderedDict([("name", (None, "pythonPlan"+"{0:%Y-%m-%d-%H-%M}".format(datetime.datetime.now()),'')),("type", (None, "NEW",'')),("platformCode", (None, "TMALL",''))\
                            ,("file", (planFileName, open(resultPath +'/'+ planFileName, 'rb').read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')) \
                            #,("file", ('WLQTESTHH001-1.xls', content,'application/vnd.ms-excel'))  \
                         #,file
                         ])
    boundary=uuid4().hex
    m = encode_multipart_formdata(files, boundary=boundary)
    print("0", m[0])
    params = {}
    #url = 'https://ross-workbench-sit.baozun.com/schedule/createSchedule'
    header = {
        'Content-Type': 'multipart/form-data;boundary={0}'.format(boundary),
        'Cookie': 'ross_token_workbench_'+environment+'='+token,
        'shopCode': shopcode
    }
    #multiple_files = [
    #    ('WLQTESTHH001-1', (None, open("WLQTESTHH001-1.xlsx", 'rb').read(), 'application/octet-stream'))]
    response = requests.post(url,
                              params=params,
                              data=m[0],
                              headers=header)

    print("1: ", response.text)
    print("2: ", response.request.body)
    print("3: ", response.request.headers)


'''https://pim-sit.baozun.com/pm/job/raw-data-row/importList/v2
'''
def getToken(env, psd):
    header = {
        'Content-Type': 'application/x-www-form-urlencoded',
        'X-Requested-With': 'XMLHttpRequest'
    }
    params = {'appkey':'ross_'+env,
              'password':psd,
              'loginName':'wsh12490',
              'domainLogin': 'false',
              'saasTenantCode':'baozun'
              }
    response = requests.post('http://ecs-uat-account.baozun.com/person/login',
                              params=params,
                              headers=header)

    #print("1: ", response.text)
    #print("2: ", response.request.body)
    #print("3: ", response.headers)
    #a = response.headers
    resHeaderDict = response.headers
    #'SESSION=984f0cf9-6b72-4695-9f2d-f49784631f54; Path=/; HttpOnly'
    setCookieStr = str(resHeaderDict['Set-Cookie'])
    reg = 'SESSION=(.+?);'
    sessionId = re.search(reg,setCookieStr).group(1)
    header2 = {'Cookie': 'SESSION='+sessionId,
               'Connection': 'keep-alive'
               }
    response2 = requests.get('http://ecs-uat-account.baozun.com/oauth/back/ross_'+env+'?callbackurl=https%3A%2F%2Fross-'+ env+ '.baozun.com%2Fgap%2Fdashboard',
                              params=params,
                              headers=header2,
                              allow_redirects=False)
    reg2 = 'code=(.+?)&'
    #sessionId2 = re.search(reg2, setCookieStr).group(1)
    #print("1: ", response2.text)
    #print("3: ", response2.headers)
    res2HeaderDict = response2.headers
    locationStr = str(res2HeaderDict['Location'])
    #print(locationStr)
    locationCode = re.search(reg2, locationStr).group(1)
    #print(locationCode)
    params3 = {'code':locationCode,
               'callbackurl':'https%3A%2F%2Fross-'+env+'.baozun.com%2Fdashboard'}
    response3 = requests.get('https://ross-auth-'+env+'.baozun.com/auth/uac/code',
                             params=params3,
                             headers=header2,
                              allow_redirects=False
                             )
    #print("1: ", response3.text)
    #print("2: ", response3.request.body)
    #print("3: ", response3.headers)
    reg3 = 'ross_token=(.+?);'
    res3HeaderDict = response3.headers
    print(str(res3HeaderDict))
    SetCookie2Str = str(res3HeaderDict['Set-Cookie'])
    ross_token = re.search(reg3, SetCookie2Str).group(1)
    #print(ross_token)
    return(ross_token)
def sanbiaobnew(huohao,cchao,environment):
    url = 'https://pim-'+ environment +'.baozun.com/pm/api/open/saveThreeStandard'
    data1 = [{
        "����ָ��": "����",
        "��������FALL": "2020/01/27",
        "������ɫ����": "����ӻ�",
        "�Ǽ���Ա": "�·���",
        "EC�ֿ��": "0",
        "�Ƿ��ѷ���": "1",
        "�㳤": "0",
        "ϴˮ��ʽ": "�뷴��ϴ�ӣ�����/����ͬ��ɫ����ϴ�ӣ�����/�������̱�ǩ",
        "��ע": "���Ž�",
        "����SKU": "313516299",
        "VAS����Ի����ƣ�": "�ѷ������ظ��۸�",
        "����&��������Ʒͼ��������": "2020/2/17",
        "Ʒ��": "POLO��",
        "�������߼�": "399",  # ��ȷ�۸�
        "��������HOL": "2020/2/14",
        "��ȫ���": "B��",
        "��Ŀ&": "test",  # ����ֶ�����
        "ʵ����Ƽ�": "399",
        "����": "Բ��",
        "ϵͳ���Ƽ�": "399",
        "������Ƽ�": "399",
        "������Ϣ�Ǽ�����": "2019/12/20",
        "������ע": "�ѷ���",
        "���Ƽ۱�ע": "Ʒ���ѻظ�",
        "��Ŀ2": "���п�",
        "��Ŀ1": "��װ",
        "����": "ɽ������",
        "��ʽ": "0",
        "�Ա�": "��ʿ",
        "���Ƴ�����Ϣ": "180/96A(M)",
        "���": huohao,   #����
        "�ѵ�������": "2019/12/31",
        "uniqueId": cchao, #ɫ��
        "���ϳɷ�": "����:��77%,������ά23%:��ñ����:��100%,������װ�β���",
        "����": "����",
        "�䳤": "����",
        "Ψһ��ʶ": cchao, #ɫ��
        "������Laydown": "0",
        "����": "",
        "����Ǽ�ʱ��": "2020-02-24 10:00:00",
        "����׼ȷ����ˣ�ʱ��": "2020-02-24 13:00:00"
    }]
    #print(data1[0])
    data = json.dumps(data1)
    headers_test = {
        'Accept': '*/*'
    }
    response = requests.post(url,data=data,headers= headers_test)
    response2 = json.loads(json.dumps(response.json()))
    a = print("�����������" + str(response2))
    return a

import json
import requests

def add_chima(huohao,cchao,environment):
    url_test = "https://pim-"+environment+".baozun.com/pm/api/open/addSizeTable/GAP"
    headers_test = {
        'Accept': '*/*'
    }
    requests_test = [{
"zipSizeSourceDataFileName":huohao, #����
"zipSizeSourceDataArticleNo":huohao,#����
"sizeType":"Regular",
 "operation": "GAP",
"dataList":[{
        "Name":"1028",
        "XS":"5.875",
        "S":"6.25",
        "M":"6.625",
        "L":"7.0",
        "XL":"7.375"
        }]
}]
    requests_test_3 = json.dumps(requests_test)
    # requests_test_3 = json.dumps(eval(requests_test))
    # print(requests_test_3)
    #print(type(requests_test_3))
    # requests_test_3 = json.dumps(requests_test)
    requestsult_test = requests.post(url_test, data=requests_test_3,headers=headers_test)
    result_test_3 = json.loads(json.dumps(requestsult_test.json()))
    a = print('�����������'+str(result_test_3))
    return a

def upload_data():
	
	'''https://ross-workbench-sit.baozun.com/schedule/createSchedule
	
	Content-Type: multipart/form-data; boundary=----WebKitFormBoundaryvXdgPDb4elLfrHfp
Cookie: ross_token_workbench_sit=
	
	------WebKitFormBoundaryvXdgPDb4elLfrHfp
Content-Disposition: form-data; name="name"

aaaa1
------WebKitFormBoundaryvXdgPDb4elLfrHfp
Content-Disposition: form-data; name="type"

NEW
------WebKitFormBoundaryvXdgPDb4elLfrHfp
Content-Disposition: form-data; name="platformCode"

TMALL
------WebKitFormBoundaryvXdgPDb4elLfrHfp
Content-Disposition: form-data; name="file"; filename="SIT data - ����.xlsx"
Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet


------WebKitFormBoundaryvXdgPDb4elLfrHfp--'''




from pymongo import MongoClient
from dateutil import parser
import random
def mdm(huohao, cchao, environment):


    
    if environment == 'sit':
        connect  = MongoClient('mongodb://u_ross_pm_service_sit:root1234@ylf-qimen-sit-mongo-db01.cloud.bz:27018,ylf-qimen-sit-mongo-db02.cloud.bz:27018,ylf-qimen-sit-mongo-db03.cloud.bz:27018/db_ross_pm_service_sit?authSource=db_ross_pm_service_sit&replicaSet=rstest-mongo4')
        database = 'db_ross_pm_service_sit'
    elif environment == 'uat':
        connect  = MongoClient('mongodb://u_ross_pm_service_uat:-6Bd%7BRJ%2Ci(FaBr6%7DS%233@10.88.27.140:27018/?authSource=db_ross_pm_service_uat')
        database = 'db_ross_pm_service_uat'
    test_db = connect[database]
    collection  = test_db['pm_raw']
    #db = client.pm_raw
    #my_set = db.pm_raw
    dateStr = "2019-05-14 01:11:11"
    myDatetime = parser.parse(dateStr)
    num = random.randint(10000, 99999)
    document = {  "sku_salesPriceHK": "",
        "sku_salesPriceTW": "",
        "sku_listPriceHK": "",
        "sku_listPriceTW": "",
        "category_extensionCode_0": "0114>1145>11450005>114500050000",
        "spu_season_attrIsSales": "",
        "spu_spuListPrice": "299.00",
        "spu_phase_attrIsSales": "",
        "sku_platformCode": "314149978",
        "category_name_0": "WOMENS>WOMENS SWEATERS>ITEMS>COTTON",
        "spu_spuSalePrice": "299.00",
        "sku_color_name": "NAVY UNIFORM",
        "sku_size_name": "XXS",
        "sku_brandCode": huohao+cchao,
        "category_attrList_catSaleProps_color": "��ɫ",
        "create_time": myDatetime,
        "sku_extensionCode": "1_000578268_HO10-11603_M03-0009",
        "brand_name": "GS",
        "sku_barCode": "1200049442895",
        "spu_spuStyleNo": huohao,
        "sku_listPrice": "299.00",
        "sku_size_attrIsSales": "",
        "sku_code": "GAP571057_1_965"+ huohao + str(num) + "_HO10-11603_M03-0009",
        "sku_size_code": "M03-0009",
        "uniqueId": "GAP571057_1_965"+ huohao + str(num) + "_HO10-11603_M03-0009",
        "brand_code": "GAP_BR000077",
        "spu_name": "V-OS BEACH CAR",
        "spu_spuGrossWeight": "11.2000",
        "spu_season_name": "Seasonal Basics 3",
        "spu_spuListPriceHK": "",
        "spu_spuListPriceTW": "",
        "mdm_sku": "1_000"+ huohao + "_HO10-11603_M03-0009",
        "spu_phase_name": "20_Y08",
        "pm_operation": "GAP",
        "spu_extensionCode": "1_000578268_HO10-11603",
        "sku_sizeDesc": "XXS",
        "BATCH": "2019-11-07.09:13:38.934",
        "sku_ext3Code": "314149978",
        "sku_color_code": "HO10-11603",
        "spu_spuSalePriceTW": "",
        "spu_spuSalePriceHK": "",
        "spu_code": "GAP_SP284653",
        "spu_season_code": "503",
        "category_attrList_catSaleProps_size": "�ߴ�",
        "spu_phase_code": "503_150",
        "brand_extensionCode": "1",
        "sku_salesPrice": "299.00",
        "sku_color_attrIsSales": "",
        "spu_styleDesc": "V-OS BEACH CAR",
        "category_code_0": "GAP_CA002060>GAP_CA002364>GAP_CA002413>GAP_CA002415",
        "pm_source": "MDM",
        "colorCode": cchao}
    #posts = db.post
    #post_id = posts.insert(post)
    one_insert  = collection.insert_one(document=document)
    print(one_insert.inserted_id)
    print(environment+"mdm�ɹ�")
    
def pdpWenAn(huohao,environment):
    path = Path.cwd()
    excels = path.glob('*.xls*')
    modelCount=0
    for excel in excels:
        filename = str(excel.name)
        if "����ҳ�İ�" in filename and "~$����ҳ�İ�" not in filename:
            #print("hello world"+filename)
            print("�ҵ�����ҳ�İ�"+"'"+filename+"'"+"���޸�ģ����")
            modelCount = modelCount + 1
            #data = xlrd.open_workbook(filename)
            data = openpyxl.load_workbook(filename)
            #book = copy(data)
            #sheet = book.get_sheet(0)
            table = data['�İ�����']
            #print(tables.size)
            
            print('���İ��������޸�ģ��')
            #nrows = table.nrows
            currColumn = 1 
            maxColumnLen = table.max_column 
            huohaoCount = 0
            while currColumn <=  maxColumnLen:
                #print(table.cell(1,currColumn).value.strip())
                if table.cell(1,currColumn).value.strip() == "����":
                    table.cell(2,currColumn).value = huohao
                    huohaoCount = huohaoCount+1
                currColumn = currColumn + 1
            if huohaoCount == 1 :
                time_stamp = '{0:%Y-%m-%d-%H-%M}'.format(datetime.datetime.now())
                data.save(str(path) +'/����ҳ'+huohao+'ʱ��'+str(time.time())+'.xlsx')
            else:
                print('û�ҵ���Ӧ�����ֶ�')
            
    if modelCount>1:
        print("�ҵ�����2������ҳ�İ�ģ�壬�����ļ�")
    elif modelCount == 0:
        print("û�ҵ�����ҳ�İ�ģ��")
    
def bitianshuxing(huohao,environment):
    path = Path.cwd()
    excels = path.glob('*.xls*')
    modelCount=0
    for excel in excels:
        filename = str(excel.name)
        if "��������ģ��" in filename and "~$��������ģ��" not in filename:
            #print("hello world"+filename)
            print("�ҵ���������ģ��"+"'"+filename+"'"+"���޸�ģ����")
            modelCount = modelCount + 1
            #data = xlrd.open_workbook(filename)
            data = openpyxl.load_workbook(filename)
            #book = copy(data)
            #sheet = book.get_sheet(0)
            table = data['spu��Ϣ']
            #print(tables.size)
            
            print('��spu��Ϣ���޸�ģ��')
            #nrows = table.nrows
            currColumn = 1 
            maxColumnLen = table.max_column 
            huohaoCount = 0
            while currColumn <=  maxColumnLen:
                #print(table.cell(1,currColumn).value.strip())
                if table.cell(1,currColumn).value.strip() == "����":
                    table.cell(2,currColumn).value = huohao
                    huohaoCount = huohaoCount+1
                currColumn = currColumn + 1
            if huohaoCount == 1 :
                time_stamp = '{0:%Y-%m-%d-%H-%M}'.format(datetime.datetime.now())
                data.save(str(path) +'/��������'+huohao+'ʱ��'+str(time.time())+'.xlsx')
            else:
                print('û�ҵ���Ӧ�����ֶ�')
    if modelCount>1:
        print("�ҵ�����2����������ģ�壬�����ļ�")
    elif modelCount == 0:
        print("û�ҵ���������ģ��")

def imageProcess(path,time_stamp,spuCodes):
    requireNum = len(spuCodes)
    resultPath = str(path) + '/' + time_stamp + '��' + str(requireNum) + '��/ԭͼ' + str(requireNum) + '��'
    sampleDir = str(path) + '/ԭͼsample'
    originImageFoloders = []
    if os.path.exists(sampleDir):
        num = 0
        while num < len(spuCodes):
            spuOriginImageFolder = resultPath + '/' + str(spuCodes[num])
            shutil.copytree(sampleDir,spuOriginImageFolder)
            num = num + 1
            originImageFoloders.append(spuOriginImageFolder)
        zipPath = str(path) + '/' + time_stamp + '��' + str(requireNum) + '��/' + str(requireNum) + '��ԭͼ.zip'
        get_zip(originImageFoloders,zipPath,resultPath)#resultPath����ɾ��ѹ���ļ�����㼶
    else:
        print('ԭͼsample�ļ��в����ڣ������˳�')
        exit()

def get_zip(files,zip_name,resultPath):
    zp=zipfile.ZipFile(zip_name,'w', zipfile.ZIP_DEFLATED)
    for file in files:
        fil = file.split('/')
        filename = fil[len(fil)-1]
        for i in os.walk(file):
            for n in i[2]:
                route = ''.join((i[0],'/',n))
                zp.write(route,route.replace(resultPath+'/',''))

    zp.close()
    print('ѹ�����')
def spuGenerate(requiredNum):
    spuCodes = []
    huohao = []
    a = 0
    while a < requiredNum:
        huohao.append("ROSSTS" + datetime.datetime.now().strftime('%m%d') + str(random.randint(100, 999)))
        spuCodes.append(huohao[a] + '-' + str(random.randint(1, 99)))
        a = a + 1
    return spuCodes, huohao

uploadYN = False
environment = 'uat'
spuCount = 1
password = '3RWnuOGFUdlzgC8HdFtSVg=='
nocc = 'false'

path = Path.cwd()
excels = path.glob('*.xls*')

spuCodes, huohao=spuGenerate(spuCount)
time_stamp = '{0:%Y%m%d%H%M}'.format(datetime.datetime.now())
planFileName = pimFileGenerate(environment,spuCount,time_stamp,spuCodes,huohao)
originImageFolders = imageProcess(path,time_stamp,spuCodes)
if uploadYN:
    token = getToken(environment,password)
    if environment == 'sit':
        planUrl = 'https://ross-workbench-sit.baozun.com'
    elif environment == 'uat':
        planUrl = 'https://ross-workbench-uat.baozun.com'
    else:
        print("unknown environment")
        exit()
    shopcode = 'abercrombiefitch'
    createSchedule(planUrl,environment,token,shopcode,time_stamp,planFileName,spuCount)
'''for excel in excels:
    filename = str(excel.name)
    if "SIT data" in filename and "~$SIT data" not in filename:
        #
        data = openpyxl.load_workbook(filename)
        table = data['Sheet1']
        maxRowLen = table.max_row
        print("����" + str(maxRowLen) + "����׼������")
        currRow = 2
        while currRow <= maxRowLen:
            mdm(table.cell(currRow,1).value,table.cell(currRow,2).value,environment)
            maihuoqingdan(table.cell(currRow,1).value,table.cell(currRow,2).value,environment)
            sanbiaobnew(table.cell(currRow,1).value,table.cell(currRow,2).value,environment)
            add_chima(table.cell(currRow,1).value,table.cell(currRow,2).value,environment)
            pdpWenAn(table.cell(currRow,1).value,environment)
            bitianshuxing(table.cell(currRow,1).value,environment)
            currRow = currRow + 1
        print("ִ�����")
'''

